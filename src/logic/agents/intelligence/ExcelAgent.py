#!/usr/bin/env python3
# Copyright 2026 PyAgent Authors
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# limitations under the License.

"""Agent specializing in structured extraction from Excel files (ExStruct pattern)."""

from __future__ import annotations
from src.core.base.version import VERSION
from src.core.base.BaseAgent import BaseAgent
import logging
from pathlib import Path
from typing import Dict, Any

__version__ = VERSION

class ExcelAgent(BaseAgent):
    """Parses Excel workbooks into structured JSON (tables, shapes, charts)."""
    
    def __init__(self, file_path: str) -> None:
        super().__init__(file_path)
        self._system_prompt = (
            "You are the Excel Specialist Agent. "
            "Your role is to extract structured data from Excel workbooks. "
            "Beyond raw cell values, you identify tables, charts, and diagrams (shapes) "
            "to provide a semantic understanding of the spreadsheet."
        )

    def extract_structured_data(self, excel_path: str, mode: str = "standard") -> dict[str, Any]:
        """Performs deep structured extraction. Supports openpyxl if available."""
        logging.info(f"ExcelAgent: Extracting data from {excel_path} in '{mode}' mode.")
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            results = {"book_name": Path(excel_path).name, "sheets": {}}
            
            for sheet in wb.worksheets:
                results["sheets"][sheet.title] = {
                    "dimensions": sheet.dimensions,
                    "merged_cells": len(sheet.merged_cells),
                    "table_candidates": [t.name for t in getattr(sheet, '_tables', [])],
                    "charts": len(getattr(sheet, '_charts', [])),
                    "max_row": sheet.max_row,
                    "max_col": sheet.max_column
                }
            return results
        except ImportError:
            logging.warning("openpyxl not found, falling back to basic metadata.")
            return {
                "book_name": Path(excel_path).name,
                "sheets": {"Sheet1": {"note": "Install openpyxl for deep parsing"}}
            }
        except Exception as e:
            return {"error": str(e)}

    def generate_markdown_summary(self, extraction_result: dict[str, Any]) -> str:
        """Converts structured Excel JSON into an AI-readable Markdown summary."""
        summary = [f"# Excel Summary: {extraction_result.get('book_name')}"]
        
        for sheet_name, data in extraction_result.get("sheets", {}).items():
            summary.append(f"## Sheet: {sheet_name}")
            summary.append(f"- **Primary Table**: {', '.join(data.get('table_candidates', []))}")
            if data.get("charts"):
                summary.append(f"- **Charts**: {len(data['charts'])} detected (Types: {', '.join(c['type'] for c in data['charts'])})")
            if data.get("shapes"):
                summary.append(f"- **Diagram Elements**: {len(data['shapes'])} shapes found.")
                
        return "\n".join(summary)

    def improve_content(self, task: str) -> str:
        """Handles Excel-related tasks."""
        if "extract" in task.lower() or ".xlsx" in task.lower():
            # In real use, we'd grab the file path from the context
            res = self.extract_structured_data("sample.xlsx")
            return self.generate_markdown_summary(res)
        return "ExcelAgent: Ready to parse spreadsheets."

if __name__ == "__main__":
    from src.core.base.utilities import create_main_function
    main = create_main_function(ExcelAgent)
    main()